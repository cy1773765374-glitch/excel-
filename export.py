# 使用提醒:
# 1. xbot包提供软件自动化、数据表格、Excel、日志、AI等功能
# 2. package包提供访问当前应用数据的功能，如获取元素、访问全局变量、获取资源文件等功能
# 3. 当此模块作为流程独立运行时执行main函数
# 4. 可视化流程中可以通过"调用模块"的指令使用此模块

import os
import re
import io
from collections import defaultdict

from xbot import print as xprint

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple


# =========================
# 安全打印：永远只传 1 个字符串参数给 xbot.print
# =========================
def xlog(msg):
    try:
        xprint(str(msg))
    except Exception:
        try:
            import builtins
            builtins.print(str(msg))
        except Exception:
            pass


# =========================
# 文件名清洗 + 防覆盖路径
# =========================
def _safe_filename(name, default="image"):
    if name is None:
        name = default
    name = str(name).strip()
    if not name:
        name = default
    return re.sub(r'[\\/:*?"<>|]+', "_", name)

def _unique_path(dir_path, filename):
    base, ext = os.path.splitext(filename)
    p = os.path.join(dir_path, filename)
    if not os.path.exists(p):
        return p
    for n in range(2, 9999):
        p2 = os.path.join(dir_path, "{}_{}{}".format(base, n, ext))
        if not os.path.exists(p2):
            return p2
    return p


# =========================
# openpyxl：解析图片 anchor 的 行/列（1-based）
# =========================
def _get_img_row_col_openpyxl(img):
    anch = getattr(img, "anchor", None)

    # OneCellAnchor / TwoCellAnchor
    if hasattr(anch, "_from"):
        try:
            r = anch._from.row + 1  # 0-based -> 1-based
            c = anch._from.col + 1
            return r, c, anch
        except Exception:
            return None, None, anch

    # "A1" 这种字符串
    if isinstance(anch, str):
        try:
            r, c = coordinate_to_tuple(anch)
            return r, c, anch
        except Exception:
            return None, None, anch

    # 兜底：某些版本字段叫 from
    fr = getattr(anch, "from", None)
    if fr is not None and hasattr(fr, "row") and hasattr(fr, "col"):
        try:
            return fr.row + 1, fr.col + 1, anch
        except Exception:
            return None, None, anch

    return None, None, anch


# =========================
# openpyxl 导出（适用于传统浮动图片 / drawing 图片）
# 返回: (exported_count, exported_rows_set)
# =========================
def _export_by_openpyxl(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug):
    ext = os.path.splitext(xlsx_path)[1].lower()
    if ext not in [".xlsx", ".xlsm"]:
        xlog("openpyxl: 仅支持 .xlsx/.xlsm，当前: {}".format(xlsx_path))
        return 0, set()

    wb = load_workbook(xlsx_path)
    if sheetName not in wb.sheetnames:
        xlog("openpyxl: 找不到工作表 {}，实际: {}".format(sheetName, wb.sheetnames))
        return 0, set()

    ws = wb[sheetName]
    imgs = list(getattr(ws, "_images", []) or [])
    xlog("openpyxl 检测到图片对象数量: {}".format(len(imgs)))

    if len(imgs) == 0:
        return 0, set()

    target_col = None
    if imgCol is not None and str(imgCol).strip() != "":
        try:
            target_col = column_index_from_string(str(imgCol).strip())
        except Exception:
            target_col = None

    # row -> [img,...]
    row2imgs = defaultdict(list)
    max_img_row = 0

    for idx, img in enumerate(imgs, start=1):
        r, c, anch = _get_img_row_col_openpyxl(img)
        if r is None:
            if int(debug) == 1:
                xlog("openpyxl IMG#{} anchor无法解析".format(idx))
            continue

        if r > max_img_row:
            max_img_row = r

        if int(debug) == 1:
            xlog("openpyxl IMG#{} row={} col={} anchType={}".format(idx, r, c, type(anch).__name__))

        if r < int(startRow):
            continue

        if target_col is None:
            row2imgs[r].append(img)
        else:
            try:
                if abs(int(c) - int(target_col)) <= int(colTolerance):
                    row2imgs[r].append(img)
            except Exception:
                row2imgs[r].append(img)

    # max_row 取更大者
    max_row = ws.max_row
    if max_img_row > max_row:
        max_row = max_img_row

    exported = 0
    exported_rows = set()

    for r in range(int(startRow), int(max_row) + 1):
        base_name = _safe_filename(ws["{}{}".format(nameCol, r)].value, default="row_{}".format(r))
        imgs_in_row = row2imgs.get(r, [])

        if not imgs_in_row:
            continue

        for k, img in enumerate(imgs_in_row, start=1):
            try:
                data = img._data()
                from PIL import Image as PILImage
                pil = PILImage.open(io.BytesIO(data))

                fmt = (pil.format or "PNG").upper()
                ext2 = "jpg" if fmt == "JPEG" else fmt.lower()

                filename = "{}_{}.{}".format(base_name, k, ext2)
                save_path = _unique_path(imgSavePath, filename)

                pil.save(save_path)
                exported += 1
                exported_rows.add(r)
                xlog("openpyxl OK: {}".format(save_path))
            except Exception as e:
                xlog("openpyxl ERR: row={}, err={}".format(r, repr(e)))

    xlog("openpyxl 导出数量: {}".format(exported))
    return exported, exported_rows


# =========================
# Excel COM：把剪贴板图片粘到临时 Chart 再 Export
# =========================
def _chart_export_from_clipboard(xl_ws, width, height, save_path):
    co = xl_ws.ChartObjects().Add(0, 0, max(10, int(width)), max(10, int(height)))
    chart = co.Chart
    chart.Paste()
    chart.Export(save_path)
    co.Delete()


# =========================
# Excel COM 导出（适用于“嵌入单元格图片 / 单元格图片类型 / openpyxl拿不全”）
# 支持：按 shapes 导出（多图/行） + fallback 单元格 CopyPicture
# skip_rows: set(row) -> auto 模式下跳过 openpyxl 已导出的行，避免重复
# 返回: exported_count
# =========================
def _export_by_com(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug, skip_rows=None):
    if skip_rows is None:
        skip_rows = set()

    try:
        import win32com.client  # type: ignore
    except Exception as e:
        xlog("COM: 缺少 pywin32，无法使用 Excel COM。err={}".format(repr(e)))
        return 0

    target_col = None
    if imgCol is not None and str(imgCol).strip() != "":
        try:
            target_col = column_index_from_string(str(imgCol).strip())
        except Exception:
            target_col = None

    name_col_idx = column_index_from_string(str(nameCol).strip())

    excel = None
    wb = None
    exported = 0

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        wb = excel.Workbooks.Open(xlsx_path)
        xl_ws = wb.Worksheets(sheetName)

        # 找 last_row（用 nameCol）
        # -4162 = xlUp
        last_row = xl_ws.Cells(xl_ws.Rows.Count, name_col_idx).End(-4162).Row

        # 1) 先按 Shapes 建立 row -> shapes（图片类型）
        pic_types = set([11, 13])  # msoLinkedPicture=11, msoPicture=13
        row2shapes = defaultdict(list)

        shapes = xl_ws.Shapes
        count_shapes = shapes.Count
        xlog("COM Shapes.Count: {}".format(count_shapes))

        for i in range(1, count_shapes + 1):
            shp = shapes.Item(i)

            try:
                stype = int(shp.Type)
            except Exception:
                continue

            if stype not in pic_types:
                continue

            try:
                tl = shp.TopLeftCell
                r = int(tl.Row)
                c = int(tl.Column)
            except Exception:
                continue

            if r < int(startRow):
                continue

            if int(debug) == 1:
                xlog("COM SHP#{} row={} col={} w={} h={}".format(i, r, c, int(shp.Width), int(shp.Height)))

            if target_col is None:
                row2shapes[r].append(shp)
            else:
                try:
                    if abs(int(c) - int(target_col)) <= int(colTolerance):
                        row2shapes[r].append(shp)
                except Exception:
                    row2shapes[r].append(shp)

        # 2) 逐行导出：优先 shapes；否则 fallback 单元格截图
        for r in range(int(startRow), int(last_row) + 1):
            if r in skip_rows:
                continue

            raw_name = xl_ws.Cells(r, name_col_idx).Value
            base_name = _safe_filename(raw_name, default="row_{}".format(r))

            shapes_in_row = row2shapes.get(r, [])

            # 2.1 行内多图：逐 shape 导出
            if shapes_in_row:
                for k, shp in enumerate(shapes_in_row, start=1):
                    try:
                        # CopyPicture(Format=2 => xlBitmap)
                        shp.CopyPicture(Format=2)
                        filename = "{}_{}.png".format(base_name, k)
                        save_path = _unique_path(imgSavePath, filename)
                        _chart_export_from_clipboard(xl_ws, shp.Width, shp.Height, save_path)
                        exported += 1
                        xlog("COM OK(shape): {}".format(save_path))
                    except Exception as e:
                        xlog("COM ERR(shape): row={}, err={}".format(r, repr(e)))
                continue

            # 2.2 fallback：导出该行 imgCol 单元格的“可视内容”
            try:
                col_idx = target_col if target_col is not None else 1
                rng = xl_ws.Cells(r, col_idx)
                # CopyPicture(Appearance=1: xlScreen, Format=2: xlBitmap)
                rng.CopyPicture(Appearance=1, Format=2)

                filename = "{}_1.png".format(base_name)
                save_path = _unique_path(imgSavePath, filename)

                _chart_export_from_clipboard(xl_ws, rng.Width, rng.Height, save_path)
                exported += 1
                xlog("COM OK(cell): {}".format(save_path))
            except Exception as e:
                if int(debug) == 1:
                    xlog("COM NOIMG row={} name={} err={}".format(r, base_name, repr(e)))

        xlog("COM 导出数量: {}".format(exported))
        return exported

    except Exception as e:
        xlog("COM 总异常: {}".format(repr(e)))
        return exported

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


# =========================
# 对外入口（做法一）
# engine:
#   - "auto": openpyxl 先导出 -> COM 只补齐 openpyxl 未导出的行
#   - "openpyxl": 只走 openpyxl
#   - "com": 只走 COM
# =========================
def export_images_by_row(xlsx_path, imgSavePath, sheetName="Sheet2",
                         nameCol="B", imgCol="A", startRow=1,
                         colTolerance=2, debug=0, engine="auto"):
    if not os.path.isfile(xlsx_path):
        xlog("xlsx_path 不存在: {}".format(xlsx_path))
        return False

    if not os.path.isdir(imgSavePath):
        xlog("imgSavePath 目录不存在: {}".format(imgSavePath))
        return False

    if engine == "openpyxl":
        try:
            c1, _ = _export_by_openpyxl(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug)
            return c1 > 0
        except Exception as e:
            xlog("openpyxl 模式异常: {}".format(repr(e)))
            return False

    if engine == "com":
        c2 = _export_by_com(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug, skip_rows=set())
        return c2 > 0

    # auto：两种都跑，但 COM 只补齐 openpyxl 未导出的行
    c1 = 0
    rows1 = set()
    try:
        c1, rows1 = _export_by_openpyxl(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug)
    except Exception as e:
        xlog("AUTO: openpyxl 异常: {}".format(repr(e)))

    c2 = 0
    try:
        c2 = _export_by_com(xlsx_path, imgSavePath, sheetName, nameCol, imgCol, startRow, colTolerance, debug, skip_rows=rows1)
    except Exception as e:
        xlog("AUTO: COM 异常: {}".format(repr(e)))

    xlog("AUTO: 完成（openpyxl={}, com={}）".format(c1, c2))
    return (c1 + c2) > 0


# =========================
# xbot 入口
# =========================
def main(args):
    try:
        xlsx_path = args.get("xlsx_path", "")
        imgSavePath = args.get("imgSavePath", "")
        sheetName = args.get("sheetName", "Sheet2")
        nameCol = args.get("nameCol", "B")
        imgCol = args.get("imgCol", "A")
        startRow = args.get("startRow", 1)
        colTolerance = args.get("colTolerance", 2)
        debug = args.get("debug", 0)
        engine = args.get("engine", "auto")  # ✅ 新增：auto/openpyxl/com

        return export_images_by_row(
            xlsx_path=xlsx_path,
            imgSavePath=imgSavePath,
            sheetName=sheetName,
            nameCol=nameCol,
            imgCol=imgCol,
            startRow=startRow,
            colTolerance=colTolerance,
            debug=debug,
            engine=engine
        )
    except Exception as e:
        xlog("main 执行异常: {}".format(repr(e)))
        return False
